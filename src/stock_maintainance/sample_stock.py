from __future__ import annotations

import json
import math
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import duckdb
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from .database import DB_PATH, connect
from .paths import REPORTS_DIR, ROOT
from .schema import quote_ident


BASE_SAMPLE_TABLES = [
    "stock_basic_info",
    "stock_company_info",
    "stock_status_history",
    "stock_daily",
    "stock_daily_basic",
    "stock_adj_factor",
    "stock_limit_price",
    "stock_moneyflow_daily",
    "margin_detail",
    "northbound_holding",
    "top_list_daily",
    "top_inst_detail",
    "financial_income_raw",
    "financial_balance_raw",
    "financial_cashflow_raw",
    "financial_indicator_raw",
    "financial_dividend_raw",
    "financial_disclosure_schedule",
    "pledge_stat",
]

DERIVED_SAMPLE_TABLES = [
    "derived_daily_spine",
    "derived_price_technical",
    "derived_volume_liquidity",
    "derived_return_momentum",
    "derived_volatility_risk",
    "derived_trading_constraint",
    "derived_valuation_size",
    "derived_financial_asof",
    "derived_financial_quality",
    "derived_financial_growth",
    "derived_capital_flow",
    "derived_sector_concept_context",
    "derived_index_market_context",
    "derived_cross_sectional",
    "derived_corporate_action",
    "derived_ownership_governance",
    "derived_composite_state",
    "stock_features_core",
    "stock_features_plus",
    "stock_features_full",
]

DATE_PRIORITY = ["trade_date", "ann_date", "first_ann_date", "end_date", "cal_date", "update_time"]


@dataclass(frozen=True)
class SampleStockResult:
    payload: dict[str, Any]
    json_path: Path
    xlsx_path: Path | None


def _table_exists(con: duckdb.DuckDBPyConnection, table_name: str) -> bool:
    return bool(
        con.execute(
            """
            SELECT count(*)
            FROM information_schema.tables
            WHERE table_name = ?
            """,
            [table_name],
        ).fetchone()[0]
    )


def _columns(con: duckdb.DuckDBPyConnection, table_name: str) -> list[str]:
    return [str(row[1]) for row in con.execute(f"PRAGMA table_info({quote_ident(table_name)})").fetchall()]


def _date_column(columns: list[str]) -> str | None:
    for name in DATE_PRIORITY:
        if name in columns:
            return name
    return None


def _clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    return value


def _rows_to_dicts(columns: list[str], rows: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    return [{col: _clean(value) for col, value in zip(columns, row)} for row in rows]


def _pick_default_stock(con: duckdb.DuckDBPyConnection) -> str:
    for table in ["derived_daily_spine", "stock_daily", "stock_basic_info"]:
        if not _table_exists(con, table):
            continue
        columns = _columns(con, table)
        if "ts_code" not in columns:
            continue
        order_sql = "ORDER BY trade_date DESC, ts_code" if "trade_date" in columns else "ORDER BY ts_code"
        row = con.execute(f"SELECT ts_code FROM {quote_ident(table)} WHERE ts_code IS NOT NULL {order_sql} LIMIT 1").fetchone()
        if row and row[0]:
            return str(row[0])
    raise ValueError("cannot infer sample ts_code; please pass --ts-code")


def _stock_profile(con: duckdb.DuckDBPyConnection, ts_code: str) -> dict[str, Any]:
    profile = {"ts_code": ts_code}
    if _table_exists(con, "stock_basic_info"):
        columns = _columns(con, "stock_basic_info")
        rows = con.execute(
            f"SELECT * FROM stock_basic_info WHERE ts_code = ? LIMIT 1",
            [ts_code],
        ).fetchall()
        if rows:
            profile.update(_rows_to_dicts(columns, rows)[0])
    return profile


def _sample_table(
    con: duckdb.DuckDBPyConnection,
    *,
    table_name: str,
    ts_code: str,
    start_date: str | None,
    end_date: str | None,
    row_limit: int,
) -> dict[str, Any]:
    if not _table_exists(con, table_name):
        return {"table": table_name, "exists": False, "date_column": None, "rows": []}
    columns = _columns(con, table_name)
    if "ts_code" not in columns:
        return {"table": table_name, "exists": True, "date_column": None, "rows": []}
    date_col = _date_column(columns)
    where = ["ts_code = ?"]
    params: list[Any] = [ts_code]
    if date_col and start_date:
        where.append(f"CAST({quote_ident(date_col)} AS DATE) >= CAST(? AS DATE)")
        params.append(start_date)
    if date_col and end_date:
        where.append(f"CAST({quote_ident(date_col)} AS DATE) <= CAST(? AS DATE)")
        params.append(end_date)
    order_sql = f"ORDER BY {quote_ident(date_col)} DESC" if date_col else ""
    rows = con.execute(
        f"""
        SELECT *
        FROM {quote_ident(table_name)}
        WHERE {" AND ".join(where)}
        {order_sql}
        LIMIT ?
        """,
        [*params, row_limit],
    ).fetchall()
    return {
        "table": table_name,
        "exists": True,
        "date_column": date_col,
        "rows": _rows_to_dicts(columns, rows),
    }


def _quality_for_table(con: duckdb.DuckDBPyConnection, table_name: str, ts_code: str) -> dict[str, Any]:
    if not _table_exists(con, table_name):
        return {"table": table_name, "exists": False}
    columns = _columns(con, table_name)
    if "ts_code" not in columns:
        return {"table": table_name, "exists": True, "has_ts_code": False}
    date_col = _date_column(columns)
    row_count = int(con.execute(f"SELECT count(*) FROM {quote_ident(table_name)} WHERE ts_code = ?", [ts_code]).fetchone()[0])
    payload: dict[str, Any] = {
        "table": table_name,
        "exists": True,
        "has_ts_code": True,
        "date_column": date_col,
        "row_count": row_count,
    }
    if date_col:
        min_date, max_date = con.execute(
            f"""
            SELECT min(CAST({quote_ident(date_col)} AS DATE)), max(CAST({quote_ident(date_col)} AS DATE))
            FROM {quote_ident(table_name)}
            WHERE ts_code = ?
            """,
            [ts_code],
        ).fetchone()
        duplicate_keys = int(
            con.execute(
                f"""
                SELECT count(*)
                FROM (
                    SELECT ts_code, {quote_ident(date_col)}, count(*) AS row_count
                    FROM {quote_ident(table_name)}
                    WHERE ts_code = ?
                    GROUP BY ts_code, {quote_ident(date_col)}
                    HAVING count(*) > 1
                )
                """,
                [ts_code],
            ).fetchone()[0]
        )
        payload.update({"min_date": _clean(min_date), "max_date": _clean(max_date), "duplicate_date_keys": duplicate_keys})
    return payload


def _safe_sheet_name(raw_name: str, used: set[str]) -> str:
    base = "".join("_" if char in "[]*/\\?:" else char for char in raw_name)[:28] or "Sheet"
    name = base
    index = 1
    while name in used:
        suffix = f"_{index}"
        name = f"{base[:31 - len(suffix)]}{suffix}"
        index += 1
    used.add(name)
    return name


def _safe_table_name(raw_name: str) -> str:
    cleaned = "".join(char if char.isalnum() or char == "_" else "_" for char in raw_name)
    if not cleaned or not cleaned[0].isalpha():
        cleaned = f"T_{cleaned}"
    return cleaned[:200]


def _write_table_sheet(ws, rows: list[dict[str, Any]], table_name: str) -> None:
    if not rows:
        rows = [{"status": "empty", "message": "No rows matched the sample filter."}]
    headers = []
    seen = set()
    for row in rows:
        for key in row:
            if key not in seen:
                headers.append(key)
                seen.add(key)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])
    _style_sheet(ws, table_name)


def _style_sheet(ws, table_name: str) -> None:
    ws.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row >= 1 and max_col >= 1:
        ref = f"A1:{get_column_letter(max_col)}{max_row}"
        table = Table(displayName=_safe_table_name(table_name), ref=ref)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        width = min(max(max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells) + 2, 10), 42)
        ws.column_dimensions[letter].width = width


def _index_rows(items: list[dict[str, Any]], sheet_names: dict[str, str]) -> list[dict[str, Any]]:
    rows = []
    for item in items:
        table = item["table"]
        rows.append(
            {
                "table": table,
                "sheet": sheet_names.get(table, ""),
                "exists": item["exists"],
                "date_column": item.get("date_column") or "",
                "sample_rows": len(item.get("rows", [])),
                "first_sample_date": _first_date(item),
                "last_sample_date": _last_date(item),
            }
        )
    return rows


def _first_date(item: dict[str, Any]) -> Any:
    date_col = item.get("date_column")
    rows = item.get("rows") or []
    if not date_col or not rows:
        return ""
    return rows[-1].get(date_col) or ""


def _last_date(item: dict[str, Any]) -> Any:
    date_col = item.get("date_column")
    rows = item.get("rows") or []
    if not date_col or not rows:
        return ""
    return rows[0].get(date_col) or ""


def _build_sample_excel(payload: dict[str, Any], output_prefix: str) -> Path:
    output_dir = ROOT / "outputs" / "phase5"
    output_dir.mkdir(parents=True, exist_ok=True)
    ts_code = payload["stock"]["ts_code"]
    safe_code = ts_code.replace(".", "_")
    output_path = output_dir / f"{output_prefix}_{safe_code}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    used = {"Summary"}
    base_sheet_names = {item["table"]: _safe_sheet_name(f"B_{item['table']}", used) for item in payload["base_tables"]}
    derived_sheet_names = {item["table"]: _safe_sheet_name(f"D_{item['table']}", used) for item in payload["derived_tables"]}

    ws.append(["Phase 5 Stock Sample Workbook"])
    ws.append([])
    summary_rows = [
        ["证券代码", payload["stock"].get("ts_code")],
        ["证券名称", payload["stock"].get("name")],
        ["市场", payload["stock"].get("market")],
        ["交易所", payload["stock"].get("exchange")],
        ["样本起始日期", payload["filters"].get("start_date")],
        ["样本结束日期", payload["filters"].get("end_date")],
        ["每表最大行数", payload["filters"].get("row_limit")],
        ["基础表数量", len(payload["base_tables"])],
        ["衍生表数量", len(payload["derived_tables"])],
        ["生成时间", payload["generated_at"]],
    ]
    for row in summary_rows:
        ws.append(row)
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = PatternFill("solid", fgColor="244062")
    ws.merge_cells("A1:H1")
    for row in range(3, 13):
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].fill = PatternFill("solid", fgColor="D9EAF7")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 32

    quality_ws = wb.create_sheet("Quality_Report")
    _write_table_sheet(quality_ws, payload["quality_report"], "QualityReportTable")

    base_index_ws = wb.create_sheet("Base_Index")
    _write_table_sheet(base_index_ws, _index_rows(payload["base_tables"], base_sheet_names), "BaseIndexTable")
    _link_index_sheet(base_index_ws)

    derived_index_ws = wb.create_sheet("Derived_Index")
    _write_table_sheet(derived_index_ws, _index_rows(payload["derived_tables"], derived_sheet_names), "DerivedIndexTable")
    _link_index_sheet(derived_index_ws)

    for item in payload["base_tables"]:
        sheet_name = base_sheet_names[item["table"]]
        table_ws = wb.create_sheet(sheet_name)
        _write_table_sheet(table_ws, item["rows"], _safe_table_name(f"{sheet_name}_Table"))
    for item in payload["derived_tables"]:
        sheet_name = derived_sheet_names[item["table"]]
        table_ws = wb.create_sheet(sheet_name)
        _write_table_sheet(table_ws, item["rows"], _safe_table_name(f"{sheet_name}_Table"))

    wb.save(output_path)
    return output_path


def _link_index_sheet(ws) -> None:
    headers = [cell.value for cell in ws[1]]
    if "sheet" not in headers:
        return
    col_idx = headers.index("sheet") + 1
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value:
            cell.hyperlink = f"#{cell.value}!A1"
            cell.style = "Hyperlink"


def sample_stock(
    *,
    ts_code: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    row_limit: int = 20,
    output_prefix: str = "phase5_sample_stock",
    build_excel: bool = True,
    db_path: Path = DB_PATH,
) -> SampleStockResult:
    with connect(db_path) as con:
        sample_ts_code = ts_code or _pick_default_stock(con)
        base_tables = [
            _sample_table(con, table_name=table, ts_code=sample_ts_code, start_date=start_date, end_date=end_date, row_limit=row_limit)
            for table in BASE_SAMPLE_TABLES
        ]
        derived_tables = [
            _sample_table(con, table_name=table, ts_code=sample_ts_code, start_date=start_date, end_date=end_date, row_limit=row_limit)
            for table in DERIVED_SAMPLE_TABLES
        ]
        quality_rows = [
            {**_quality_for_table(con, table, sample_ts_code), "category": "base"}
            for table in BASE_SAMPLE_TABLES
        ] + [
            {**_quality_for_table(con, table, sample_ts_code), "category": "derived"}
            for table in DERIVED_SAMPLE_TABLES
        ]
        payload = {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "stock": _stock_profile(con, sample_ts_code),
            "filters": {"start_date": start_date, "end_date": end_date, "row_limit": row_limit},
            "base_tables": base_tables,
            "derived_tables": derived_tables,
            "quality_report": quality_rows,
        }

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    safe_code = sample_ts_code.replace(".", "_")
    json_path = REPORTS_DIR / f"{output_prefix}_{safe_code}.json"
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, allow_nan=False) + "\n", encoding="utf-8")
    xlsx_path = None
    if build_excel:
        xlsx_path = _build_sample_excel(payload, output_prefix)
    return SampleStockResult(payload=payload, json_path=json_path, xlsx_path=xlsx_path)
